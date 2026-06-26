import threading
import time
import logging
import os
import statistics
from datetime import datetime
from collections import defaultdict
from typing import Dict, List

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("[DriftLogger] openpyxl not installed — run: pip3 install openpyxl")

HEADER_FONT  = Font(name="Arial", bold=True, color="58A6FF", size=10)
DATA_FONT    = Font(name="Arial", size=10, color="E6EDF3")
SUMMARY_FONT = Font(name="Arial", bold=True, color="3FB950", size=10)
TITLE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=12)

def _hfill(color="1A1A2E"):
    return PatternFill("solid", start_color=color, fgColor=color)

def _border():
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)

class DriftLogger:

    def __init__(self, output_dir: str = "/root",
                 interval_s: float = 60.0,
                 max_samples: int = 50000):
        self.output_dir  = output_dir
        self.interval_s  = interval_s
        self.max_samples = max_samples
        self._start_time = time.time()
        self._lock       = threading.Lock()
        self._running    = False
        self._data: Dict[str, List[dict]] = defaultdict(list)
        self._roles: Dict[str, str] = {}
        self.filepath    = os.path.join(output_dir, "clock_drift.xlsx")

    def start(self):
        if not OPENPYXL_AVAILABLE:
            logger.warning("[DriftLogger] openpyxl missing — logging disabled")
            return
        self._running = True
        threading.Thread(target=self._write_loop, daemon=True,
                         name="drift-logger").start()
        logger.info("[DriftLogger] Started. Updating %s every %ds",
                    self.filepath, self.interval_s)

    def stop(self):
        self._running = False
        self.flush()

    def record(self, node_id: str, role: str, offset_ms: float):
        if not OPENPYXL_AVAILABLE:
            return
        now = time.time()
        with self._lock:
            self._roles[node_id] = role
            samples = self._data[node_id]
            samples.append({
                "ts":        datetime.fromtimestamp(now),
                "elapsed_s": round(now - self._start_time, 2),
                "offset_ms": round(offset_ms, 4),
            })
            if len(samples) > self.max_samples:
                self._data[node_id] = samples[-self.max_samples:]

    def flush(self):
        if not OPENPYXL_AVAILABLE:
            return
        with self._lock:
            data  = {nid: list(s) for nid, s in self._data.items() if s}
            roles = dict(self._roles)
        if not data:
            return
        self._write_excel(data, roles)

    def get_last_file(self):
        return self.filepath

    def _write_loop(self):
        while self._running:
            time.sleep(self.interval_s)
            if self._running:
                self.flush()

    def _write_excel(self, data: dict, roles: dict):
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            all_stats = {}
            for node_id in sorted(data.keys()):
                samples = data[node_id]
                role    = roles.get(node_id, "unknown")
                stats   = self._write_node_sheet(wb, node_id, role, samples)
                all_stats[node_id] = stats

            self._write_summary_sheet(wb, all_stats)
            wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)
            wb.save(self.filepath)

            total = sum(len(s) for s in data.values())
            logger.info("[DriftLogger] Updated %s — %d total samples",
                        self.filepath, total)
        except Exception as e:
            logger.error("[DriftLogger] Write error: %s", e)

    def _write_node_sheet(self, wb, node_id, role, samples):
        ws = wb.create_sheet(title=node_id[:31])
        ws.sheet_properties.tabColor = "58A6FF"

        ws.merge_cells("A1:D1")
        ws["A1"] = f"Clock Drift — {node_id}  ({role.upper()})"
        ws["A1"].font      = TITLE_FONT
        ws["A1"].fill      = _hfill("0D1117")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for col, h in enumerate(["Timestamp","Elapsed (s)","Offset (ms)","Node ID"], 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = HEADER_FONT; cell.fill = _hfill(); cell.border = _border()
            cell.alignment = Alignment(horizontal="center")

        offsets = []
        for i, s in enumerate(samples):
            row  = i + 3
            fill = _hfill("161B22") if i % 2 == 0 else _hfill("0D1117")
            for col, v in enumerate([s["ts"], s["elapsed_s"], s["offset_ms"], node_id], 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = DATA_FONT; cell.fill = fill; cell.border = _border()
                if col == 1:
                    cell.number_format = "YYYY-MM-DD HH:MM:SS"
                    cell.alignment = Alignment(horizontal="left")
                elif col in (2, 3):
                    cell.number_format = "0.0000"
                    cell.alignment = Alignment(horizontal="right")
            offsets.append(s["offset_ms"])

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.freeze_panes = "A3"
        return self._stats(offsets, node_id, role)

    def _stats(self, offsets, node_id, role):
        if not offsets: return {}
        s = sorted(offsets); n = len(s)
        return {
            "node_id": node_id, "role": role, "count": n,
            "mean": round(statistics.mean(offsets), 4),
            "std":  round(statistics.stdev(offsets) if n > 1 else 0.0, 4),
            "min":  round(min(offsets), 4),
            "p50":  round(s[int(0.50 * n)], 4),
            "p95":  round(s[min(int(0.95 * n), n-1)], 4),
            "p99":  round(s[min(int(0.99 * n), n-1)], 4),
            "max":  round(max(offsets), 4),
        }

    def _write_summary_sheet(self, wb, all_stats):
        ws = wb.create_sheet(title="Summary")
        ws.sheet_properties.tabColor = "3FB950"

        ws.merge_cells("A1:J1")
        ws["A1"] = "Virtual TSN — Clock Drift Summary"
        ws["A1"].font = TITLE_FONT; ws["A1"].fill = _hfill("0D1117")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.merge_cells("A2:J2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Arial", italic=True, color="8B949E", size=9)

        headers = ["Node ID","Role","Samples","Mean (ms)","Std (ms)",
                   "Min (ms)","P50 (ms)","P95 (ms)","P99 (ms)","Max (ms)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT; cell.fill = _hfill()
            cell.alignment = Alignment(horizontal="center"); cell.border = _border()

        for i, (nid, st) in enumerate(sorted(all_stats.items())):
            if not st: continue
            row  = i + 4
            fill = _hfill("161B22") if i % 2 == 0 else _hfill("0D1117")
            vals = [st["node_id"], st["role"], st["count"],
                    st["mean"], st["std"], st["min"],
                    st["p50"], st["p95"], st["p99"], st["max"]]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = SUMMARY_FONT if col <= 2 else DATA_FONT
                cell.fill = fill; cell.border = _border()
                if col >= 4:
                    cell.number_format = "0.0000"
                    cell.alignment = Alignment(horizontal="right")

        for col, w in enumerate([16,12,10,12,12,12,12,12,12,12], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A4"
